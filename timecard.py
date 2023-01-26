#!/usr/bin/python3
# SPDX-FileCopyrightText: 2023 Ryuji Hirano
# SPDX-License-Identifier: BSD-3-Clause

import os
import openpyxl
import datetime

#Excelファイルパス指定
book = openpyxl.load_workbook('dates.xlsx')
sheet = book.active

#日時取得
dt = datetime.datetime.now()

#配列
dta = [dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second]

print(len(dta))

i = 1

while True:
    if sheet.cell(row=i,column=1).value is None:
        break
    i+=1

rows = i
clm = 1

for thing in dta:
    sheet.cell(row=rows,column=clm).value = thing
    clm+=1

book.save('dates.xlsx')
